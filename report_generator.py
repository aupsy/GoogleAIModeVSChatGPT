"""
Report Generator Module
Generates comprehensive Excel reports with charts and statistical analysis.
"""

from datetime import datetime
from typing import Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

import analyzer
import data_manager


def create_excel_report(output_file: str = None) -> str:
    """
    Generate comprehensive Excel report.

    Args:
        output_file: Output filename (default: auto-generated with timestamp)

    Returns:
        Path to generated file
    """
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'AI_Evaluation_Report_{timestamp}.xlsx'

    # Generate analysis
    print("Generating analysis...")
    analysis = analyzer.generate_full_analysis()

    if 'error' in analysis:
        print(f"Error: {analysis['error']}")
        return None

    df = analyzer.load_analysis_data()

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    print("Creating Summary sheet...")
    create_summary_sheet(wb, analysis)

    print("Creating By Category sheet...")
    create_category_sheet(wb, analysis)

    print("Creating By Quality sheet...")
    create_quality_sheet(wb, analysis)

    print("Creating By Intent Clarity sheet...")
    create_intent_clarity_sheet(wb, analysis)

    print("Creating By Web Search sheet...")
    create_web_search_sheet(wb, analysis)

    print("Creating Raw Data sheet...")
    create_raw_data_sheet(wb, df)

    print("Creating Individual Queries sheet...")
    create_individual_queries_sheet(wb, df)

    # Save workbook
    print(f"Saving report to {output_file}...")
    wb.save(output_file)

    print(f"✓ Report generated successfully: {output_file}")
    return output_file


def create_summary_sheet(wb: Workbook, analysis: Dict):
    """Create summary sheet with overall statistics."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = 'AI Evaluation Report: ChatGPT vs Google AI Mode'
    ws['A1'].font = Font(size=16, bold=True)

    # Metadata
    row = 3
    ws[f'A{row}'] = 'Report Generated:'
    ws[f'B{row}'] = analysis['metadata']['generated_at']
    row += 1
    ws[f'A{row}'] = 'Total Queries Analyzed:'
    ws[f'B{row}'] = analysis['metadata']['total_queries']

    # Key Insights
    row += 3
    ws[f'A{row}'] = 'Key Insights'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1

    insights = analyzer.generate_insights(analysis)
    for insight in insights:
        ws[f'A{row}'] = f'• {insight}'
        row += 1

    # Summary Statistics Table
    row += 2
    ws[f'A{row}'] = 'Summary Statistics'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1

    # Header row
    headers = ['Metric', 'ChatGPT Mean', 'ChatGPT Std', 'Google AI Mean', 'Google AI Std', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    row += 1

    # Data rows
    summary = analysis['summary_stats']
    for metric in ['relevance', 'completeness', 'source_quality']:
        chatgpt_mean = summary['chatgpt'].get(metric, {}).get('mean', 0)
        chatgpt_std = summary['chatgpt'].get(metric, {}).get('std', 0)
        google_mean = summary['google'].get(metric, {}).get('mean', 0)
        google_std = summary['google'].get(metric, {}).get('std', 0)
        diff = google_mean - chatgpt_mean

        ws.cell(row, 1, metric.capitalize())
        ws.cell(row, 2, chatgpt_mean)
        ws.cell(row, 3, chatgpt_std)
        ws.cell(row, 4, google_mean)
        ws.cell(row, 5, google_std)
        ws.cell(row, 6, diff)

        # Color code difference
        diff_cell = ws.cell(row, 6)
        if diff > 0:
            diff_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif diff < 0:
            diff_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        row += 1

    # Statistical Tests
    row += 2
    ws[f'A{row}'] = 'Statistical Significance Tests'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1

    # Test results table
    headers = ['Metric', 'T-Statistic', 'P-Value', 'Significant?', "Cohen's d", 'Effect Size']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    row += 1

    tests = analysis['statistical_tests']
    for metric, test_result in tests.items():
        if 'error' in test_result:
            continue

        ws.cell(row, 1, metric.capitalize())
        ws.cell(row, 2, test_result['t_statistic'])
        ws.cell(row, 3, test_result['p_value'])
        ws.cell(row, 4, 'Yes' if test_result['significant'] else 'No')
        ws.cell(row, 5, test_result['cohens_d'])
        ws.cell(row, 6, test_result['interpretation'].capitalize())

        # Highlight significant results
        if test_result['significant']:
            for col in range(1, 7):
                ws.cell(row, col).font = Font(bold=True)

        row += 1

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def create_category_sheet(wb: Workbook, analysis: Dict):
    """Create sheet with performance by category."""
    ws = wb.create_sheet("By Category")

    # Title
    ws['A1'] = 'Performance by Query Category'
    ws['A1'].font = Font(size=14, bold=True)

    row = 3
    by_category = analysis['by_category']

    # Create table
    headers = ['Category', 'Count', 'ChatGPT Rel', 'ChatGPT Comp', 'ChatGPT Source',
               'Google Rel', 'Google Comp', 'Google Source']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    row += 1

    data_start_row = row
    for category, stats in by_category.items():
        ws.cell(row, 1, category)
        ws.cell(row, 2, stats['count'])

        chatgpt = stats.get('chatgpt', {})
        ws.cell(row, 3, chatgpt.get('relevance', ''))
        ws.cell(row, 4, chatgpt.get('completeness', ''))
        ws.cell(row, 5, chatgpt.get('source_quality', ''))

        google = stats.get('google', {})
        ws.cell(row, 6, google.get('relevance', ''))
        ws.cell(row, 7, google.get('completeness', ''))
        ws.cell(row, 8, google.get('source_quality', ''))

        row += 1

    # Add chart
    chart = BarChart()
    chart.title = "Average Scores by Category"
    chart.y_axis.title = "Score (1-5)"
    chart.x_axis.title = "Category"

    categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=row-1)
    chatgpt_data = Reference(ws, min_col=3, max_col=5, min_row=data_start_row-1, max_row=row-1)
    google_data = Reference(ws, min_col=6, max_col=8, min_row=data_start_row-1, max_row=row-1)

    chart.add_data(chatgpt_data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, f"A{row+2}")

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 30)


def create_quality_sheet(wb: Workbook, analysis: Dict):
    """Create sheet with performance by query quality."""
    ws = wb.create_sheet("By Quality")

    # Title
    ws['A1'] = 'Performance by Query Quality'
    ws['A1'].font = Font(size=14, bold=True)

    row = 3
    by_quality = analysis['by_quality']

    # Create table
    headers = ['Quality Level', 'Count', 'ChatGPT Avg', 'ChatGPT Intent %',
               'Google Avg', 'Google Intent %']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    row += 1

    # Quality level order for logical display
    quality_order = ['Well-formed', 'Poorly-formed', 'Ambiguous', 'Typos/Informal', 'Time-sensitive ambiguous']

    data_start_row = row
    for quality in quality_order:
        if quality not in by_quality:
            continue

        stats = by_quality[quality]
        ws.cell(row, 1, quality)
        ws.cell(row, 2, stats['count'])

        # Calculate average across metrics
        chatgpt = stats.get('chatgpt', {})
        chatgpt_avg = sum([v for k, v in chatgpt.items()
                          if k in ['relevance', 'completeness', 'source_quality']]) / 3 if chatgpt else 0
        ws.cell(row, 3, round(chatgpt_avg, 2))
        ws.cell(row, 4, chatgpt.get('intent_rate', ''))

        google = stats.get('google', {})
        google_avg = sum([v for k, v in google.items()
                         if k in ['relevance', 'completeness', 'source_quality']]) / 3 if google else 0
        ws.cell(row, 5, round(google_avg, 2))
        ws.cell(row, 6, google.get('intent_rate', ''))

        row += 1

    # Add chart
    chart = LineChart()
    chart.title = "Performance Drop-off by Query Quality"
    chart.y_axis.title = "Average Score"
    chart.x_axis.title = "Quality Level"

    categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=row-1)
    chatgpt_data = Reference(ws, min_col=3, min_row=data_start_row-1, max_row=row-1)
    google_data = Reference(ws, min_col=5, min_row=data_start_row-1, max_row=row-1)

    chart.add_data(chatgpt_data, titles_from_data=True)
    chart.add_data(google_data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, f"A{row+2}")

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 30)


def create_intent_clarity_sheet(wb: Workbook, analysis: Dict):
    """Create sheet with performance by intent clarity."""
    ws = wb.create_sheet("By Intent Clarity")

    # Title
    ws['A1'] = 'Performance by Intent Clarity Level'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = 'Testing the hypothesis: Google AI Mode has better intent understanding'
    ws['A2'].font = Font(italic=True)

    row = 4
    by_clarity = analysis['by_intent_clarity']

    # Create table
    headers = ['Intent Clarity', 'Count', 'ChatGPT Avg', 'ChatGPT Intent %',
               'Google Avg', 'Google Intent %', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    row += 1

    # Clarity order
    clarity_order = ['High', 'Medium', 'Low', 'Very Low']

    data_start_row = row
    for clarity in clarity_order:
        if clarity not in by_clarity:
            continue

        stats = by_clarity[clarity]
        ws.cell(row, 1, clarity)
        ws.cell(row, 2, stats['count'])

        # Calculate average
        chatgpt = stats.get('chatgpt', {})
        chatgpt_avg = sum([v for k, v in chatgpt.items()
                          if k in ['relevance', 'completeness', 'source_quality']]) / 3 if chatgpt else 0
        ws.cell(row, 3, round(chatgpt_avg, 2))
        ws.cell(row, 4, chatgpt.get('intent_rate', ''))

        google = stats.get('google', {})
        google_avg = sum([v for k, v in google.items()
                         if k in ['relevance', 'completeness', 'source_quality']]) / 3 if google else 0
        ws.cell(row, 5, round(google_avg, 2))
        ws.cell(row, 6, google.get('intent_rate', ''))

        diff = google_avg - chatgpt_avg
        ws.cell(row, 7, round(diff, 2))

        # Highlight differences
        diff_cell = ws.cell(row, 7)
        if diff > 0:
            diff_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif diff < 0:
            diff_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        row += 1

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 30)


def create_web_search_sheet(wb: Workbook, analysis: Dict):
    """Create sheet with ChatGPT performance by web search usage."""
    ws = wb.create_sheet("By Web Search")

    # Title
    ws['A1'] = 'ChatGPT Performance: Web Search Impact Analysis'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = 'Comparing ChatGPT performance when web search is used vs not used'
    ws['A2'].font = Font(italic=True)

    row = 4
    by_search = analysis.get('by_web_search', {})

    if 'error' in by_search:
        ws.cell(row, 1, by_search['error'])
        ws.cell(row, 1).font = Font(italic=True)
        return

    # Create table
    headers = ['Web Search Status', 'Count', 'Relevance', 'Completeness',
               'Source Quality', 'Overall Avg', 'Intent Rate %']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    row += 1

    # With web search
    with_search = by_search.get('with_web_search', {})
    if with_search.get('count', 0) > 0:
        ws.cell(row, 1, 'With Web Search')
        ws.cell(row, 2, with_search['count'])

        chatgpt = with_search.get('chatgpt', {})
        ws.cell(row, 3, chatgpt.get('relevance', ''))
        ws.cell(row, 4, chatgpt.get('completeness', ''))
        ws.cell(row, 5, chatgpt.get('source_quality', ''))

        # Overall average
        metrics = [chatgpt.get('relevance', 0), chatgpt.get('completeness', 0),
                  chatgpt.get('source_quality', 0)]
        avg = sum(metrics) / len([m for m in metrics if m > 0]) if any(metrics) else 0
        ws.cell(row, 6, round(avg, 2))
        ws.cell(row, 7, chatgpt.get('intent_rate', ''))
        row += 1

    # Without web search
    without_search = by_search.get('without_web_search', {})
    if without_search.get('count', 0) > 0:
        ws.cell(row, 1, 'Without Web Search')
        ws.cell(row, 2, without_search['count'])

        chatgpt = without_search.get('chatgpt', {})
        ws.cell(row, 3, chatgpt.get('relevance', ''))
        ws.cell(row, 4, chatgpt.get('completeness', ''))
        ws.cell(row, 5, chatgpt.get('source_quality', ''))

        # Overall average
        metrics = [chatgpt.get('relevance', 0), chatgpt.get('completeness', 0),
                  chatgpt.get('source_quality', 0)]
        avg = sum(metrics) / len([m for m in metrics if m > 0]) if any(metrics) else 0
        ws.cell(row, 6, round(avg, 2))
        ws.cell(row, 7, chatgpt.get('intent_rate', ''))
        row += 1

    # Add comparison section
    row += 2
    ws.cell(row, 1, 'Impact of Web Search (Difference)')
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    comparison = by_search.get('comparison', {})
    if comparison:
        headers = ['Metric', 'Difference', 'Percent Change']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        row += 1

        for metric in ['relevance', 'completeness', 'source_quality']:
            if metric in comparison:
                ws.cell(row, 1, metric.title())
                ws.cell(row, 2, comparison[metric]['difference'])
                ws.cell(row, 3, f"{comparison[metric]['percent_change']}%")

                # Highlight positive/negative
                diff_val = comparison[metric]['difference']
                for col in [2, 3]:
                    cell = ws.cell(row, col)
                    if diff_val > 0:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif diff_val < 0:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                row += 1

    # Add insights
    row += 2
    ws.cell(row, 1, 'Key Insights:')
    ws.cell(row, 1).font = Font(bold=True)
    row += 1

    if with_search.get('count', 0) > 0 and without_search.get('count', 0) > 0:
        if comparison:
            ws.cell(row, 1, f"• Web search was used in {with_search['count']} out of "
                           f"{with_search['count'] + without_search['count']} queries")
            row += 1

            overall_impact = sum([comparison[m]['difference']
                                for m in ['relevance', 'completeness', 'source_quality']
                                if m in comparison]) / 3
            if overall_impact > 0.2:
                ws.cell(row, 1, "• Web search significantly IMPROVED ChatGPT performance")
            elif overall_impact < -0.2:
                ws.cell(row, 1, "• Web search DECREASED ChatGPT performance")
            else:
                ws.cell(row, 1, "• Web search had minimal impact on ChatGPT performance")
            row += 1
    else:
        ws.cell(row, 1, "• Insufficient data for web search comparison")
        row += 1

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 35)


def create_raw_data_sheet(wb: Workbook, df: pd.DataFrame):
    """Create sheet with raw data."""
    ws = wb.create_sheet("Raw Data")

    # Add DataFrame to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)

    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def create_individual_queries_sheet(wb: Workbook, df: pd.DataFrame):
    """Create sheet with individual query comparisons."""
    ws = wb.create_sheet("Individual Queries")

    row = 1
    ws[f'A{row}'] = 'Individual Query Comparisons'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 2

    # Get full responses from results
    results = data_manager.load_results()

    for _, query_row in df.iterrows():
        query_id = str(query_row['query_id'])

        # Query header
        ws[f'A{row}'] = f"Query {query_id}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = query_row['query']
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = f"Category: {query_row['category']} | Quality: {query_row['quality']} | Intent Clarity: {query_row['intent_clarity']}"
        ws[f'A{row}'].font = Font(italic=True)
        row += 1

        # Scores header
        headers = ['Platform', 'Relevance', 'Completeness', 'Source Quality', 'Intent?']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = Font(bold=True)
        row += 1

        # ChatGPT scores
        ws.cell(row, 1, 'ChatGPT')
        ws.cell(row, 2, query_row.get('chatgpt_relevance', ''))
        ws.cell(row, 3, query_row.get('chatgpt_completeness', ''))
        ws.cell(row, 4, query_row.get('chatgpt_source_quality', ''))
        ws.cell(row, 5, 'Yes' if query_row.get('chatgpt_intent_understood') else 'No')
        row += 1

        # Google AI scores
        ws.cell(row, 1, 'Google AI')
        ws.cell(row, 2, query_row.get('google_relevance', ''))
        ws.cell(row, 3, query_row.get('google_completeness', ''))
        ws.cell(row, 4, query_row.get('google_source_quality', ''))
        ws.cell(row, 5, 'Yes' if query_row.get('google_intent_understood') else 'No')
        row += 2

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 60)


if __name__ == "__main__":
    print("Generating Excel Report...\n")
    output = create_excel_report()

    if output:
        print(f"\nSuccess! Report saved to: {output}")
    else:
        print("\nFailed to generate report")
